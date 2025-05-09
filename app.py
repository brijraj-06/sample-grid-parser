import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

# Style definitions
bold_font = Font(bold=True)
italic_font = Font(italic=True)
wrap_alignment = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def generate_composition_table(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "COMPOSITION TABLE FORMAT"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 55

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "1. COMPOSITION TABLE FORMAT"
    ws['A1'].font = bold_font
    ws.merge_cells('A2:E2')
    ws['A2'] = "Each 50g contains:"

    ws.append(["#", "English Transliterated Name (Botanical Name)/ हिंदी नाम", 
               "Part Used Full Form", "Quantity", "Proof Of Concept"])
    for col in range(1, 6):
        c = ws.cell(row=3, column=col)
        c.font = bold_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border

    index = 1
    row_idx = 4
    for _, row in df.iterrows():
        name = str(row.get("Name", "")).strip()
        botanical = str(row.get("Botanical Name", "")).strip()
        hindi = str(row.get("हिंदी नाम", "")).strip()

        # Section header like "Amla Pishti:"
        if name.endswith(":") and botanical == "":
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
            c = ws.cell(row=row_idx, column=2, value=name)
            c.font = bold_font
            c.alignment = wrap_alignment
            c.border = thin_border
            row_idx += 1
            continue

        # Create combined name
        full_name = f"{name} ({botanical})/ {hindi}".replace("(nan)", "").replace("()", "").replace("(/", "(").strip(" /")

        quantity = row.get("Quantity", "")
        unit = row.get("Unit", "")
        quantity_str = f"{int(quantity) if str(quantity).replace('.', '', 1).isdigit() and float(quantity).is_integer() else quantity} {unit}".strip()

        # Fill row
        ws.cell(row=row_idx, column=1, value=index)
        ws.cell(row=row_idx, column=2, value=full_name)
        ws.cell(row=row_idx, column=3, value=row.get("Part Used Full Form", ""))
        ws.cell(row=row_idx, column=4, value=quantity_str)
        ws.cell(row=row_idx, column=5, value=row.get("Proof Of Concept", ""))

        for col in range(1, 6):
            c = ws.cell(row=row_idx, column=col)
            c.alignment = wrap_alignment
            c.border = thin_border

        index += 1
        row_idx += 1

    # Add final note
    ws.merge_cells(start_row=row_idx + 1, start_column=1, end_row=row_idx + 1, end_column=5)
    ws.cell(row=row_idx + 1, column=1, value="*Official Substitute")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_paragraph_excel(title, subtitle, lines):
    wb = Workbook()
    ws = wb.active
    ws.title = title.replace(" ", "_")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 150

    row = 1
    ws.cell(row=row, column=1, value=title).font = bold_font
    row += 1
    ws.cell(row=row, column=1, value=subtitle).font = italic_font
    row += 2

    for line in lines:
        ws.cell(row=row, column=1, value=line).alignment = wrap_alignment
        row += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Streamlit app
st.title("Master Grid → 3 Output Generator")

uploaded_file = st.file_uploader("Upload the Master Excel file", type=["xlsx"])
if uploaded_file:
    df = pd.read_excel(uploaded_file)

    required_columns = ["Name", "Botanical Name", "हिंदी नाम", "Part Used Full Form", "Quantity", "Proof Of Concept", "Unit"]
    if not all(col in df.columns for col in required_columns):
        st.error("Uploaded file must contain the required columns: Name, Botanical Name, हिंदी नाम, Part Used Full Form, Quantity, Proof Of Concept, Unit")
    else:
        st.success("File uploaded and validated.")

        comp_file = generate_composition_table(df)

        # Paragraphs (unchanged for now)
        hindi_lines = [
            "Each 50g contains:",
            "Amla Pishti:",
            "Abhrak Bhasma (RT)/ अभ्रक भस्म 1g; Bel (Aegle marmelos)/ बेल (St. Bk.), Choti Elaichi (Elettaria cardamomum)/ छोटी इलायची (Sd.) each 1.75mg.",
            "Kwath Dravya (Coarse Powders Of):",
            "Giloy (Tinospora cordifolia)/ गिलोय (St.) 20mg; Kakoli (Withania somnifera*)/ अश्वगंधा (Rt.) 10mg; Munnaka (Vitis vinifera)/ मुनक्का (Fr.), Badi Kateri (Solanum indicum)/ बड़ी कटेरी (Pl.) each 5mg. Permitted Additives QS.",
            "*Official Substitute"
        ]
        hindi_file = generate_paragraph_excel("PARAGRAPH FORMAT (ENGLISH-HINDI MIX)",
                                              "English Transliterated Name (Botanical Name)/ हिंदी नाम (Part Used Short Form) Qty.",
                                              hindi_lines)

        eng_lines = [
            "Each 50g contains:",
            "Amla Pishti:",
            "Abhrak Bhasma (RT) 1g; Bel (Aegle marmelos) (St. Bk.), Choti Elaichi (Elettaria cardamomum) (Sd.) each 1.75mg.",
            "Kwath Dravya (Coarse Powders Of):",
            "Giloy (Tinospora cordifolia) (St.) 20mg; Kakoli (Withania somnifera*) (Rt.) 10mg; Munnaka (Vitis vinifera) (Fr.), Badi Kateri (Solanum indicum) (Pl.) each 5mg. Permitted Additives QS.",
            "*Official Substitute"
        ]
        eng_file = generate_paragraph_excel("PARAGRAPH FORMAT (ENGLISH ONLY)",
                                            "English Transliterated Name (Botanical Name) (Part Used Short Form) Qty.",
                                            eng_lines)

        st.download_button("Download Composition Table Format", comp_file, file_name="COMPOSITION_TABLE_FORMAT.xlsx")
        st.download_button("Download Paragraph Format (English-Hindi Mix)", hindi_file, file_name="PARAGRAPH_FORMAT_ENGLISH_HINDI_MIX.xlsx")
        st.download_button("Download Paragraph Format (English Only)", eng_file, file_name="PARAGRAPH_FORMAT_ENGLISH_ONLY.xlsx")
